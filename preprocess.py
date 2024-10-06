import re
from concurrent.futures import ThreadPoolExecutor

import emoji
import emoji_vietnamese
import openpyxl
import csv
import inflect
from underthesea import word_tokenize
from urllib.parse import urlparse

raw_comments_file = './comments.xlsx'
output_csv = './ratings_comments.csv'
#
# dataframe = openpyxl.load_workbook(filename)
# sheet = dataframe.active

ratingByComments = {}
ratingLabel = {
    5: [0, 0, 0, 0, 1],
    4: [0, 0, 0, 1, 0],
    3: [0, 0, 1, 0, 0],
    2: [0, 1, 0, 0, 0],
    1: [1, 0, 0, 0, 0]
}


#
# with open(output_csv, 'w', newline='', encoding='utf-8') as file:
#     writer = csv.writer(file)
#     writer.writerow(['Rating', 'Comment'])
#     for comment, rating in ratingByComments.items():
#         writer.writerow([rating, comment])


def get_raw_data(filename):
    rating_by_comments = {}
    dataframe = openpyxl.load_workbook(filename)
    sheet = dataframe.active
    for row in sheet.iter_rows(2, sheet.max_row):
        ratingByComments[row[1].value] = row[0].value
    return rating_by_comments

def process_comment(comment, rating):
    r = ratingLabel[int(rating)]
    p = Preprocessor(comment)
    p.preprocess()
    return [r, p.text, p.preprocess_to_list()]

def write_to_output(filename):
    with open(filename, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(['Rating', 'Comment', 'Comment Array'])
        for comment, rating in ratingByComments.items():
            r = ratingLabel[int(rating)]
            p = Preprocessor(comment)
            p.preprocess()
            writer.writerow([r, p.text, p.preprocess_to_list()])
        # with ThreadPoolExecutor() as executor:
        #     results = list(executor.map(lambda item: process_comment(item[0], item[1]), ratingByComments.items()))
        # for result in results:
        #     writer.writerow(result)

def get_stopwords():
    stopwords = set()
    with open("vietnamese-stopwords-dash.txt", "r", newline='', encoding='utf-8') as file:
        for line in file:
            stopwords.add(line.strip())
    return stopwords


def validate_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except AttributeError:
        return False

class Preprocessor():
    def __init__(self, text):
        self.text = text

    def text_lowercase(self):
        self.text = self.text.lower()

    def remove_numbers(self):
        self.text = re.sub(r'\d+', '', self.text)

    def standardize_data(self):
        text = re.sub(r"[\.,\?]+$-", "", self.text)
        # Remove all . , " ... in sentences
        text = text.replace(",", " ").replace(".", " ") \
            .replace(";", " ").replace("“", " ") \
            .replace(":", " ").replace("”", " ") \
            .replace('"', " ").replace("'", " ") \
            .replace("!", " ").replace("?", " ") \
            .replace("-", " ").replace("?", " ")

        self.text = text

    def remove_emails(self):
        words = []
        for word in self.text.split():
            valid = re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', word)
            if not valid:
                words.append(word)
        self.text = ' '.join(words)

    def remove_urls(self):
        words = [word for word in self.text.split() if not validate_url(word)]
        self.text = ' '.join(words)

    def convert_emojis(self):
        self.text = emoji_vietnamese.demojize(string=self.text)

    def remove_whitespace(self):
        self.text = ' '.join(self.text.split())

    def tokenizer(self):
        self.text = str(word_tokenize(self.text, format="text"))

    def remove_stopwords(self):
        stopwords = get_stopwords()
        tmp_words = [word for word in self.text.split() if word not in stopwords]
        self.text = ' '.join(tmp_words)

    def preprocess_to_list(self):
        self.preprocess()
        return self.text.split()

    def preprocess(self):
        if self.text is None:
            self.text = ''
            return
        self.remove_emails()
        self.remove_urls()
        self.convert_emojis()
        self.text_lowercase()
        self.remove_numbers()
        self.standardize_data()
        self.tokenizer()
        self.remove_stopwords()
        self.remove_whitespace()


def main():
    get_raw_data(raw_comments_file)
    write_to_output(output_csv)

if __name__=='__main__':
    main()







